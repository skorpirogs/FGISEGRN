import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time
import tkinter as tk
from tkinter import filedialog
import pyautogui
import pyperclip
from builtins import ValueError, IndexError
from datetime import datetime

def start_reading():
    now = datetime.now()
    current_time = now.strftime("%H:%M:%S")

    end_time = datetime.now()
    print(current_time, " - Загрузка программы")
    end_time = datetime.now()
    print(current_time, " - Выберите файл для получения данных о помещениях")

    def choose_file():
        root = tk.Tk()
        root.withdraw()
        file_path = filedialog.askopenfilename(initialdir = "/", title = "Select file", filetypes = (("Excel files", "*.xlsx"), ("all files", "*.*")))
        return file_path

    file_path = choose_file()
    wb = openpyxl.load_workbook(file_path)
    sheet = wb["Sheet1"]

    driver = webdriver.Chrome()
    driver.get("https://rosreestr.gov.ru/wps/portal/p/cc_present/EGRN_1")
    time.sleep(3)

    #chrome.exe --disable-features=PageLoadMetrics
    #____________________________________________________________________________________
    #Вход в систему
    #code - ///-///-///-///-///

    #в настоящее время сайт ФГИС ЕГРН доступен только при подтверждении своих рисков. В связи с этим в скрипт добавлено прожатие данной кнопки.
    try:
        end_time = datetime.now()
        print(current_time, " - Поиск кнопки Касперского")
        search_button = driver.find_element(By.LINK_TEXT, "Я понимаю риск, но хочу продолжить")
        search_button.click()
        time.sleep(2)
        pyautogui.press('enter', interval=0.2)
        time.sleep(1)
        end_time = datetime.now()
        print(end_time, " - Защита пройдена, продолжаю работу")
        
    except:
        end_time = datetime.now()
        print(end_time, " - Кнопка Касперского отсутствует, продолжаю работу")

    search_buttons = driver.find_elements(By.CLASS_NAME, "v-textfield")
    second_element = search_buttons[0]
    second_element.click()
    time.sleep(0.5)

    #ввод кода
    #1 часть Вашего ключа
    pyautogui.write('///')
    time.sleep(0.5)
    #2 часть Вашего ключа
    pyautogui.write('///')
    time.sleep(0.5)
    #3 часть Вашего ключа
    pyautogui.write('///')
    time.sleep(0.5)
    #4 часть Вашего ключа
    pyautogui.write('///')
    time.sleep(0.5)
    #5 часть Вашего ключа
    pyautogui.write('///')
    time.sleep(0.5)

    #Подготовка скрипта к запуску цикла
    #кнопка войти
    search_buttons = driver.find_elements(By.CLASS_NAME, "v-button-caption")
    second_element = search_buttons[0]
    second_element.click()
    time.sleep(2)

    #кнопка поиск объектов недвижимости
    search_buttons = driver.find_elements(By.CLASS_NAME, "v-button-caption")
    second_element = search_buttons[0]
    second_element.click()
    time.sleep(0.5)

    #поле кадастрового номера
    search_buttons = driver.find_elements(By.CSS_SELECTOR, "input.v-textfield.v-textfield-prompt")
    second_element = search_buttons[0]
    second_element.click()
    time.sleep(1)

    #вставить кадастр
    pyautogui.write('77:07:0000000:10338')
    time.sleep(0.5)

    #раскрывающееся меню регионов
    search_buttons = driver.find_elements(By.CLASS_NAME, "v-filterselect-button")
    second_element = search_buttons[0]
    second_element.click()
    time.sleep(2)

    #Москва
    for i in range(31):    
        pyautogui.press('down')
    pyautogui.press('enter', interval=0.2)
    time.sleep(1)
    pyautogui.press('enter', interval=0.2)
    time.sleep(8)

    #____________________________________________________________________________________
    #ЗАПУСК ЦИКЛА
    file_counter = 1
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    # new_sheet['A1'] = 'Кадастровый номер'
    # new_sheet['B1'] = 'Полный адрес объекта'
    # new_sheet['C1'] = 'Тип объекта'
    # new_sheet['D1'] = 'Площадь'
    # new_sheet['E1'] = 'Назначение здания или помещения'
    
    for i in range(1, sheet.max_row + 1):
        while True:
            try:
                end_time = datetime.now()
                print(end_time, " - Загрузка цикла")
                #кнопка назад
                search_buttons = driver.find_elements(By.CLASS_NAME, "v-button-caption")
                second_element = search_buttons[4]
                second_element.click()
                time.sleep(2)
                
                #поле кадастрового номера
                search_buttons = driver.find_elements(By.CSS_SELECTOR, "input.v-textfield")
                second_element = search_buttons[0]
                second_element.clear()
                time.sleep(1)

                #вставка данных из строк файла Microsoft Excel
                value = sheet.cell(row=i, column=1).value
                second_element.send_keys(value)
                time.sleep(2)
                pyautogui.press('enter', interval=0.2)
                time.sleep(25)
                end_time = datetime.now()
                print(end_time, " - Запуск считывания таблицы")
                
                table = driver.find_element(By.CSS_SELECTOR, "table.v-table-table")
                rows = table.find_elements(By.TAG_NAME, "tr")

                # Создаем новый документ Microsoft Excel
                new_workbook = openpyxl.Workbook()
                new_sheet = new_workbook.active
                new_sheet.title = "Моя таблица"

                # Проходим по всем строкам таблицы Microsoft Excel
                for i, row in enumerate(rows):
                    cells = row.find_elements(By.TAG_NAME, "td")
                    # Проходим по всем ячейкам строки Microsoft Excel
                    for j, cell in enumerate(cells):
                        # Получаем текст ячейки и записываем его в новый документ Microsoft Excel
                        cell_text = cell.text
                        new_sheet.cell(row=i+1, column=j+1, value=cell_text)
                        

                # Сохраняем документ Excel
                filename = f"C:\AAA{file_counter}.xlsx"
                new_workbook.save(filename)
                file_counter += 1
                end_time = datetime.now()
                print(end_time, f" - Файл {filename} сохранен, продолжаю работу")
                
            except Exception as e:
            
                end_time = datetime.now()
                print(end_time, f" - Ошибка при обработке сайта, перезагружаюсь")
                driver.quit()
                time.sleep(2)
                driver = webdriver.Chrome()
                driver.get("https://rosreestr.gov.ru/wps/portal/p/cc_present/EGRN_1")
                time.sleep(3)
                
                try:
                    end_time = datetime.now()
                    print(end_time, " - Поиск кнопки Касперского")
                    search_button = driver.find_element(By.LINK_TEXT, "Я понимаю риск, но хочу продолжить")
                    search_button.click()
                    time.sleep(2)
                    pyautogui.press('enter', interval=0.2)
                    time.sleep(1)
                    end_time = datetime.now()
                    print(end_time, " - Защита пройдена, продолжаю работу")
            
                except:
                    end_time = datetime.now()
                    print(end_time, " - Кнопка Касперского отсутствует, продолжаю работу")
                
                search_buttons = driver.find_elements(By.CLASS_NAME, "v-textfield")
                second_element = search_buttons[0]
                second_element.click()
                time.sleep(0.5)

                #ввод кода
                pyautogui.write('///')
                time.sleep(0.5)
                pyautogui.write('///')
                time.sleep(0.5)
                pyautogui.write('///')
                time.sleep(0.5)
                pyautogui.write('///')
                time.sleep(0.5)
                pyautogui.write('///')
                time.sleep(0.5)
                
                #кнопка войти
                search_buttons = driver.find_elements(By.CLASS_NAME, "v-button-caption")
                second_element = search_buttons[0]
                second_element.click()
                time.sleep(2)

                #кнопка поиск объектов недвижимости
                search_buttons = driver.find_elements(By.CLASS_NAME, "v-button-caption")
                second_element = search_buttons[0]
                second_element.click()
                time.sleep(0.5)


                #поле кадастрового номера
                search_buttons = driver.find_elements(By.CSS_SELECTOR, "input.v-textfield.v-textfield-prompt")
                second_element = search_buttons[0]
                second_element.click()
                time.sleep(1)

                #вставить кадастр
                pyautogui.write('77:07:0000000:10338')
                time.sleep(0.5)


                #раскрывающееся меню регионов
                search_buttons = driver.find_elements(By.CLASS_NAME, "v-filterselect-button")
                second_element = search_buttons[0]
                second_element.click()
                time.sleep(2)


                #Москва
                for i in range(31):    
                    pyautogui.press('down')
                pyautogui.press('enter', interval=0.2)
                time.sleep(1)
                pyautogui.press('enter', interval=0.2)
                time.sleep(8)
                end_time = datetime.now()
                print(end_time, " - Возвращаюсь к циклу")
                
                continue
    pass

def stop_reading():
    driver.quit()

root = tk.Tk()

start_button = tk.Button(root, text="Запустить считывание", command=start_reading)
start_button.pack()

stop_button = tk.Button(root, text="Остановить считывание", command=stop_reading)
stop_button.pack()

driver = webdriver.Chrome()

root.mainloop()
